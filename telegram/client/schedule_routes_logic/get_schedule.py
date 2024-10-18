from aiogram import Router, F
from aiogram.filters.callback_data import CallbackQuery
from aiogram.fsm.context import FSMContext
from aiogram.types import Message, FSInputFile
from aiogram.utils.keyboard import InlineKeyboardBuilder, InlineKeyboardButton
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta

from ...servicies.data_manager.rnis_manager import get_route_data, get_schedule_data, get_bus_stop_data
from .. import forms, callbacks, validators
from . import messages

router = Router()
img_path = 'telegram/img/schedule_logo.jpg'

ITEMS_PER_PAGE = 7
FILL_GREEN = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Зелёная заливка
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") 
FILL_LIGHT_BLUE = PatternFill(start_color="BDD7EA", end_color="BDD7EA", fill_type="solid")
FILL_LIGHT_PURPLE = PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def get_routes_keyboard(routes_data, page: int = 0):
    builder = InlineKeyboardBuilder()
    
    start_idx = page * ITEMS_PER_PAGE
    end_idx = start_idx + ITEMS_PER_PAGE
    page_data = routes_data[start_idx:end_idx]
    
    for route in page_data:
        builder.button(
            text=route['title'],
            callback_data=callbacks.RouteInfoCallback(
                route_uuid=route['uuid']
            ).pack()
        )

    builder.adjust(1)
    
    nav_buttons = []
    if page > 0:
        nav_buttons.append(InlineKeyboardButton(
            text='⬅️', callback_data=f"prev_route_page_{page}"
        ))
    
    if end_idx < len(routes_data):
        nav_buttons.append(InlineKeyboardButton(
            text='➡️', callback_data=f"next_route_page_{page}"
        ))
    
    if nav_buttons:
        builder.row(*nav_buttons)
        
    return builder.as_markup()


def configure_excel_bytes(data):
    wb = Workbook()
    ws = wb.active

    ws.add_image(Image(img_path), 'B2')

    ws.merge_cells('H2:J2')
    ws['H2'] = 'Маршрут:'
    ws['H2'].font = Font(name='Times New Roman', size=24, bold=True)
    ws['H2'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('K2:W2')
    ws['K2'] = data['route_name']
    ws['K2'].font = Font(name='Times New Roman', size=24, bold=True)
    ws['K2'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('H3:K3')
    ws['H3'] = 'Номер маршрута:'
    ws['H3'].font = Font(name='Times New Roman', size=24, bold=True)
    ws['H3'].alignment = Alignment(horizontal='left', vertical='center')

    ws['L3'] = int(data['route_number'])
    ws['L3'].font = Font(name='Times New Roman', size=24, bold=True)
    ws['L3'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('B6:E6')
    ws['B6'] = 'Расписание актуально на:'
    ws['B6'].font = Font(name='Times New Roman', size=14, bold=True)
    ws['B6'].alignment = Alignment(horizontal='center', vertical='center')

    schedule_days = {
        "Пн": data.get('monday', False),
        "Вт": data.get('tuesday', False),
        "Ср": data.get('wednesday', False),
        "Чт": data.get('thursday', False),
        "Пт": data.get('friday', False),
        "Сб": data.get('saturday', False),
        "Вс": data.get('sunday', False),
        "Праздничные дни": data.get('holiday', False),
    }

    for i, item in enumerate(schedule_days.items(), start=2):
        cell = ws.cell(row=7, column=i, value=item[0])
        cell.font = Font(name='Times New Roman', size=14, bold=True)
        if item[0] == 'Праздничные дни':
            ws.merge_cells(f'{get_column_letter(i)}7:{get_column_letter(i+2)}7')
        cell.fill = FILL_GREEN if item[1] else FILL_RED
        cell.border = THIN_BORDER

    ws.merge_cells('B10:D10')
    ws['B10'] = 'Направление'
    ws['B10'].font = Font(name='Times New Roman', size=16, bold=True)
    ws['B10'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('E10:G10')
    ws['E10'] = 'Название остановки'
    ws['E10'].font = Font(name='Times New Roman', size=16, bold=True)
    ws['E10'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['H10'] = 'Время прибытия'

    row_index = 11
    max_arrived_count_times = 0

    for key, variants in data['variants'].items():
        direction = 'Прямое' if key == 'production_forward' else 'Обратное'

        for _, variant_data in variants.items():
            start_times = variant_data['start_times']

            for _, bus_stop_info in variant_data['bus_stops'].items():
                iter_index = 0
                stop_name = bus_stop_info['stop_name']
                prev_stop_time = []

                for interval in bus_stop_info['intervals']:
                    stop_time = datetime.strptime(start_times[iter_index], "%H:%M") 
                    stop_time += timedelta(minutes=interval)
                    prev_stop_time.append(stop_time.strftime("%H:%M"))
                    iter_index += 1
                
                start_times = prev_stop_time

                if len(start_times) > max_arrived_count_times:
                    max_arrived_count_times = len(start_times)

                column_data = [direction, stop_name] + start_times
                for index, data in enumerate(column_data, start=2):
                    if index == 2:
                        cell = ws.cell(row=row_index, column=index, value=data)
                        ws.merge_cells(f'{get_column_letter(index)}{row_index}:{get_column_letter(index + 2)}{row_index}')
                        if direction == 'Прямое':
                            cell.fill = FILL_LIGHT_BLUE
                        else:
                            cell.fill = FILL_LIGHT_PURPLE
                        
                    if index == 3:
                        cell = ws.cell(row=row_index, column=index+2, value=data)
                        ws.merge_cells(f'{get_column_letter(index)}{row_index}:{get_column_letter(index + 1)}{row_index}')
                    
                    if index > 3:
                        cell = ws.cell(row=row_index, column=index+4, value=data)
        
                row_index += 1

    max_arrived_count_times += 5
    end_column = get_column_letter(max_arrived_count_times)

    ws.merge_cells(f'H10:{end_column}10')
    ws['H10'].font = Font(name='Times New Roman', size=16, bold=True)
    ws['H10'].alignment = Alignment(horizontal='center', vertical='center')

    file_path = 'static/schedule.xlsx'
    wb.save(file_path)

    return file_path


async def get_parsed_schedule_data(schedule_data):
    response = {
        'route_name': schedule_data[0]['route_name'],
        'route_number': schedule_data[0]['route_number'],
        'monday': schedule_data[0]['monday'],
        'tuesday': schedule_data[0]['tuesday'],
        'wednesday': schedule_data[0]['wednesday'],
        'thursday': schedule_data[0]['thursday'],
        'friday': schedule_data[0]['friday'],
        'saturday': schedule_data[0]['saturday'],
        'sunday': schedule_data[0]['sunday'],
        'holiday': schedule_data[0]['holiday'],
        'variants': {
            'production_forward': {},
            'production_reverse': {}
        }
    }

    for schedule in schedule_data:
        for turn in schedule['turns']:
            for run in turn['runs']:
                if run['type'] in ['production_reverse', 'production_forward']:
                    variant_direction = run['type']
                    variant_uuid = run['route_variant_uuid']
                    start_time = run['start_time']

                    response['variants'][variant_direction][variant_uuid] = response['variants'][variant_direction].get(variant_uuid, {
                        'start_times': [],
                        'bus_stops': {}
                    })
                    response['variants'][variant_direction][variant_uuid]['start_times'].append(start_time)

                    for map_data in run['production_interval_map']:
                        stop_uuid = map_data['uuid']
                        interval = map_data['interval']
                        stop_name = await get_bus_stop_data(stop_uuid)

                        response['variants'][variant_direction][variant_uuid]['bus_stops'][stop_uuid] = response['variants'][variant_direction][variant_uuid]['bus_stops'].get(stop_uuid, {           
                            'stop_name': stop_name,
                            'intervals': [],
                        })

                        response['variants'][variant_direction][variant_uuid]['bus_stops'][stop_uuid]['intervals'].append(interval)

    return response


@router.callback_query(forms.ScheduleRouteForm.get_route, callbacks.RouteInputCallback.filter())
async def input_route_data(callback_query: CallbackQuery, callback_data: callbacks.RouteInputCallback, state: FSMContext):
    input_type = callback_data.action

    if input_type == 'number':
        await state.set_state(forms.RouteInfoImputForm.input_number)
        mess = messages.input_route_number_message

    if input_type == 'name':
        await state.set_state(forms.RouteInfoImputForm.input_name)
        mess = messages.input_route_name_message

    data = await state.get_data()
    print(data)

    await callback_query.message.edit_text(mess)


@router.message(forms.RouteInfoImputForm.input_number, F.text)
async def input_route_number(message: Message, state: FSMContext):
    data = await state.get_data()
    validator = validators.RouteInputNumberValueValidator(message.text.lower())
    valid = await validator.is_valid()

    if not valid:
        await message.answer(validator.error)
        return
    
    try:
        data['route_info']['route_number'] = message.text
        route_data = await get_route_data(data)
        if len(route_data) != 0:
            data['route_info']['route_request_items'] = route_data

            await state.set_data(data)
            await state.set_state(forms.RouteInfoImputForm.select_route)

            keyboard = get_routes_keyboard(route_data)
            mess = messages.input_route_message

            await message.answer(mess, reply_markup=keyboard)
        else:
            mess = messages.empty_routes_message
            await message.answer(mess)
    except:
        mess = messages.empty_routes_message
        await message.answer(mess)
    

@router.callback_query(lambda c: c.data.startswith('prev_route_page_') or c.data.startswith('next_route_page_'))
async def handle_route_pagination(callback_query: CallbackQuery, state: FSMContext):
    page = int(callback_query.data.split('_')[-1])
    routes_data = await state.get_data()
    if 'prev_route_page' in callback_query.data:
        new_page = page - 1
    elif 'next_route_page' in callback_query.data:
        new_page = page + 1

    keyboard = get_routes_keyboard(routes_data['route_info']['route_request_items'], new_page)

    await callback_query.message.edit_reply_markup(reply_markup=keyboard)
    await callback_query.answer()


@router.message(forms.RouteInfoImputForm.input_name, F.text)
async def input_route_name(message: Message, state: FSMContext):
    data = await state.get_data()
    data['route_info']['route_name'] = message.text
    route_data = await get_route_data(data)
    if len(route_data) != 0:
        data['route_info']['route_request_items'] = route_data

        await state.set_data(data)
        await state.set_state(forms.RouteInfoImputForm.select_route)

        keyboard = get_routes_keyboard(route_data)
        mess = messages.input_route_message

        await message.answer(mess, reply_markup=keyboard)
    else:
        mess = messages.empty_routes_message
        await message.answer(mess)


@router.callback_query(forms.RouteInfoImputForm.select_route, callbacks.RouteInfoCallback.filter())
async def get_schedule_route_data(callback_query: CallbackQuery, callback_data: callbacks.RouteInfoCallback, state: FSMContext):
    data = await state.get_data()

    route = data['route_info']['route_name'] if data['route_info'].get('route_name', None) else data['route_info']['route_number']
    municipality_name = data['route_info']['municipality_name']
    mess = messages.waiting_schedule_message.format(route=route, municipality_name=municipality_name)

    await callback_query.message.edit_text(mess)

    data['route_info']['route_uuid'] = callback_data.route_uuid

    schedule_data = await get_schedule_data(data)
    parsed_schedule_data = await get_parsed_schedule_data(schedule_data)

    xslsx_bytes = configure_excel_bytes(parsed_schedule_data)

    await callback_query.message.answer_document(
                    document=FSInputFile(xslsx_bytes, filename=f'{route}-{municipality_name}.xlsx'),
                    caption="Расписание",
                    disable_notification=True
                )
